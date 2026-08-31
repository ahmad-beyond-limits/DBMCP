import csv
import io
import json
import logging
from typing import Any, Dict, List, Optional, Tuple

from docx import Document
import openpyxl
from pypdf import PdfReader

from app.anonymisation.pii_detector import PIIDetector

logger = logging.getLogger(__name__)


class ContentExtractor:
    @classmethod
    async def extract(
        cls, content: bytes, file_type: str, filename: str = "document"
    ) -> Tuple[str, Optional[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Extracts plain text and structured data from uploaded files.
        Directly preserves document content for AI agent querying.
        """
        ft = file_type.upper()
        plain_text = ""
        structured_data = None

        if ft == "TXT":
            plain_text = content.decode("utf-8", errors="replace")

        elif ft == "PDF":
            reader = PdfReader(io.BytesIO(content))
            pages = []
            for i, page in enumerate(reader.pages):
                try:
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(text.strip())
                except Exception:
                    continue
            plain_text = "\n\n".join(pages) if pages else "[PDF Document content ready for AI query]"

        elif ft == "DOCX":
            try:
                doc = Document(io.BytesIO(content))
                paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
                plain_text = "\n\n".join(paragraphs)
            except Exception:
                plain_text = "[DOCX Document content ready for AI query]"

        elif ft == "CSV":
            plain_text, structured_data = cls._extract_csv(content)

        elif ft == "JSON":
            plain_text, structured_data = cls._extract_json(content)

        elif ft in ["XLSX", "XLS"]:
            plain_text, structured_data = cls._extract_xlsx(content)

        elif ft in ["IMAGE", "PNG", "JPG", "JPEG", "WEBP", "GIF", "SVG"]:
            plain_text, structured_data = cls._extract_image(content, filename)

        else:
            plain_text = content.decode("utf-8", errors="replace")

        # Detect PII entities across extracted plain text
        detected = PIIDetector.detect_entities(plain_text)
        return plain_text, structured_data, detected

    @classmethod
    def _extract_csv(cls, content: bytes) -> Tuple[str, Dict[str, Any]]:
        text_stream = io.StringIO(content.decode("utf-8", errors="replace"))
        reader = csv.DictReader(text_stream)
        rows = []
        columns = reader.fieldnames or []
        schema = {}
        for row in reader:
            rows.append(row)
            if len(rows) <= 500:
                for k, v in row.items():
                    if k not in schema:
                        try:
                            float(v)
                            schema[k] = "number"
                        except (ValueError, TypeError):
                            schema[k] = "string"

        lines = [",".join(columns)]
        for r in rows[:100]:
            lines.append(",".join([str(r.get(c, "")) for c in columns]))
        plain_text = "\n".join(lines)

        structured_data = {
            "columns": list(columns),
            "schema": schema,
            "row_count": len(rows),
            "rows": rows,
            "raw_matrix": [list(columns)] + [[r.get(c, "") for c in columns] for r in rows],
            "header_row_index": 0,
            "table_detected": len(columns) > 0,
        }
        return plain_text, structured_data

    @classmethod
    def _extract_json(cls, content: bytes) -> Tuple[str, Dict[str, Any]]:
        raw_text = content.decode("utf-8", errors="replace")
        try:
            parsed = json.loads(raw_text)
            if isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], dict):
                columns = list(parsed[0].keys())
                schema = {k: type(parsed[0][k]).__name__ for k in columns}
                structured_data = {
                    "columns": columns,
                    "schema": schema,
                    "row_count": len(parsed),
                    "rows": parsed,
                    "raw_matrix": [columns] + [[r.get(c, "") for c in columns] for r in parsed],
                    "header_row_index": 0,
                    "table_detected": True,
                }
            else:
                structured_data = {"raw": parsed, "table_detected": False}
        except Exception:
            structured_data = {"raw": raw_text, "table_detected": False}
        return raw_text, structured_data

    @classmethod
    def _extract_xlsx(cls, content: bytes) -> Tuple[str, Dict[str, Any]]:
        """
        Extracts tabular data across all sheets of an Excel (.xlsx / .xls) workbook.
        If structured tables are detected, extracts columns, schemas, and dictionary rows.
        If unstructured text, gracefully returns raw text content.
        """
        plain_text_sections = []
        sheets_data: Dict[str, Any] = {}
        primary_columns: List[str] = []
        primary_schema: Dict[str, str] = {}
        primary_rows: List[Dict[str, Any]] = []
        primary_raw_matrix: List[List[str]] = []
        primary_header_idx = 0

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                raw_rows = list(sheet.iter_rows(values_only=True))

                # Filter empty rows
                non_empty_rows = [
                    [str(c).strip() if c is not None else "" for c in r]
                    for r in raw_rows
                    if any(c is not None and str(c).strip() != "" for c in r)
                ]

                if not non_empty_rows:
                    continue

                if len(non_empty_rows) >= 2:
                    # Smart Header Detection: locate the row with the most distinct populated column headers
                    best_header_idx = 0
                    max_cols = 0
                    for idx, r in enumerate(non_empty_rows[:8]):
                        non_empty_cols = [c for c in r if c and str(c).strip()]
                        if len(non_empty_cols) > max_cols:
                            max_cols = len(non_empty_cols)
                            best_header_idx = idx

                    raw_headers = non_empty_rows[best_header_idx]
                    headers = []
                    for i, h in enumerate(raw_headers):
                        h_clean = str(h).strip() if h else f"Column_{i + 1}"
                        headers.append(h_clean if h_clean else f"Column_{i + 1}")

                    sheet_rows = []
                    sheet_schema: Dict[str, str] = {}

                    for r in non_empty_rows[best_header_idx + 1:]:
                        row_dict = {}
                        for i, col_name in enumerate(headers):
                            val = r[i] if i < len(r) else ""
                            row_dict[col_name] = val
                            if col_name not in sheet_schema and val != "":
                                try:
                                    float(val)
                                    sheet_schema[col_name] = "number"
                                except (ValueError, TypeError):
                                    sheet_schema[col_name] = "string"
                        sheet_rows.append(row_dict)

                    sheets_data[sheet_name] = {
                        "columns": headers,
                        "schema": sheet_schema,
                        "row_count": len(sheet_rows),
                        "rows": sheet_rows,
                        "raw_matrix": non_empty_rows,
                        "header_row_index": best_header_idx,
                        "table_detected": len(headers) > 0 and len(sheet_rows) > 0,
                    }

                    if not primary_columns:
                        primary_columns = headers
                        primary_schema = sheet_schema
                        primary_rows = sheet_rows
                        primary_raw_matrix = non_empty_rows
                        primary_header_idx = best_header_idx

                    # Build text representation for this sheet
                    lines = [f"--- Sheet: {sheet_name} ---", ",".join(headers)]
                    for r in sheet_rows[:100]:
                        lines.append(",".join([str(r.get(c, "")) for c in headers]))
                    plain_text_sections.append("\n".join(lines))
                else:
                    # Single line or unstructured text in sheet
                    text_lines = [f"--- Sheet: {sheet_name} ---"] + [",".join(r) for r in non_empty_rows]
                    plain_text_sections.append("\n".join(text_lines))

            plain_text = "\n\n".join(plain_text_sections) if plain_text_sections else "[Empty Excel Workbook]"

            structured_data = {
                "columns": primary_columns,
                "schema": primary_schema,
                "row_count": len(primary_rows),
                "rows": primary_rows,
                "raw_matrix": primary_raw_matrix,
                "header_row_index": primary_header_idx,
                "sheets": sheets_data,
                "sheet_names": list(sheets_data.keys()),
                "table_detected": len(primary_columns) > 0,
            }
            return plain_text, structured_data

        except Exception as e:
            logger.error(f"Error parsing Excel workbook: {e}", exc_info=True)
            fallback_text = "[Excel document content uploaded - raw binary ready]"
            return fallback_text, {"raw": fallback_text, "table_detected": False}

    @classmethod
    def _extract_image(cls, content: bytes, filename: str) -> Tuple[str, Dict[str, Any]]:
        """Extract metadata, dimensions, color modes, and structured info from image files."""
        fn_lower = filename.lower()
        
        # 1. Handle SVG Vector Graphics
        if fn_lower.endswith(".svg") or content.strip().startswith(b"<svg") or b"<svg" in content[:200]:
            try:
                svg_text = content.decode("utf-8", errors="replace")
                # Extract text content if present in <text> tags
                import re
                text_matches = re.findall(r"<text[^>]*>([^<]+)</text>", svg_text, re.IGNORECASE)
                extracted_words = " ".join(text_matches).strip()
                
                plain_text = f"[SVG VECTOR IMAGE: {filename}]\n"
                if extracted_words:
                    plain_text += f"Extracted Text/Labels: {extracted_words}\n"
                plain_text += f"File Size: {round(len(content)/1024, 2)} KB\nFormat: Scalable Vector Graphics (SVG)"

                structured_data = {
                    "is_image": True,
                    "is_vector": True,
                    "format": "SVG",
                    "filename": filename,
                    "size_kb": round(len(content) / 1024, 2),
                    "extracted_text": extracted_words if extracted_words else None,
                }
                return plain_text, structured_data
            except Exception as e:
                logger.warning(f"Error reading SVG content: {e}")

        # 2. Handle Binary Raster Images (PNG, JPG, WEBP, GIF)
        try:
            from PIL import Image
            img = Image.open(io.BytesIO(content))
            width, height = img.size
            img_format = img.format or "IMAGE"
            mode = img.mode
            aspect_ratio = round(width / height, 3) if height > 0 else 1.0
            size_kb = round(len(content) / 1024, 2)

            plain_text = (
                f"[IMAGE RESOURCE: {filename}]\n"
                f"Format: {img_format}\n"
                f"Resolution: {width} x {height} pixels (Aspect Ratio: {aspect_ratio})\n"
                f"Color Mode: {mode}\n"
                f"File Size: {size_kb} KB\n"
                f"Description: Visual image asset indexed for policy-governed MCP tool execution and reference."
            )

            structured_data = {
                "is_image": True,
                "is_vector": False,
                "filename": filename,
                "format": img_format,
                "width": width,
                "height": height,
                "aspect_ratio": aspect_ratio,
                "mode": mode,
                "size_kb": size_kb,
            }
            return plain_text, structured_data

        except Exception as e:
            logger.warning(f"Error parsing image file {filename}: {e}")
            plain_text = f"[IMAGE FILE: {filename} ({round(len(content)/1024, 2)} KB)]"
            return plain_text, {
                "is_image": True,
                "filename": filename,
                "size_kb": round(len(content) / 1024, 2),
                "error": str(e),
            }
